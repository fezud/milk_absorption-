import os
from openpyxl import Workbook, load_workbook
from scipy.optimize import curve_fit
import numpy as np

directory = os.getcwd()
wb = load_workbook(f'{directory}\\kappa_concentration.xlsx')

ws = wb.active

def linear(x, k):
    return k*x

concentrations = np.array([1/20, 1/25, 1/30, 1/50])

wavelengths = []
kappas = []

for row in ws.iter_rows(min_row=2):
    wavelength = float(str(row[0].value).replace(',', '.'))
    kappa_values = np.array([float(str(raw_cell_data.value).replace(',', '.')) if float(str(raw_cell_data.value).replace(',', '.')) > 0 else 0 for raw_cell_data in row[9:]])

    pars, _ = curve_fit(f=linear, xdata=concentrations, ydata=kappa_values)
    print(wavelength, pars[0])
    wavelengths.append(wavelength)
    kappas.append(pars[0])

output_wb = Workbook()
output_ws = output_wb.active

output_ws.title = "Beer Kappa"

for i in range(len(wavelengths)):
    output_ws.append((wavelengths[i], kappas[i]))

output_wb.save('Output.xlsx')